from openpyxl import Workbook
wb = Workbook()
#inside the wb --> which sheet we need to work with
#grab the active worksheet
ws = wb.active
#active sheet  --  sheet that is open OR mention the sheet_name
ws['A1'] = 42
ws.append([1,2,3])
import datetime
ws['A2'] = datetime.datetime.now()
wb.save("sample1.xlsx")